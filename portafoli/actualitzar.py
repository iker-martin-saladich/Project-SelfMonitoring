#!/usr/bin/env python3
from pathlib import Path
from openpyxl import load_workbook
import re

BASE = Path(__file__).parent

def cel(ws, row, col):
    v = ws.cell(row=row, column=col).value
    if v is None:
        return ""
    if isinstance(v, float) and v == int(v):
        return int(v)
    return v

def es_num(v):
    try:
        int(str(v).strip())
        return True
    except Exception:
        return False

def files_dades(ws):
    return [r for r in range(5, ws.max_row + 1) if es_num(ws.cell(r, 1).value)]

def to_pct(v):
    try:
        return float(v)
    except Exception:
        return 0.0

def progress_global(ws, c):
    vals = [to_pct(cel(ws, r, c)) for r in files_dades(ws)]
    return round(sum(vals) / len(vals), 1) if vals else 0

def nom_full(wb, opcions):
    for n in opcions:
        if n in wb.sheetnames:
            return wb[n]
    return wb[wb.sheetnames[0]]

C_GREEN  = "#4A7C59"; C_GREEN_BG  = "#E8F2EB"
C_ORANGE = "#A07820"; C_ORANGE_BG = "#FEF6E0"
C_GRAY   = "#8C7B68"; C_GRAY_BG   = "#F0EEEB"
C_RED    = "#8B3A3A"; C_RED_BG    = "#FDEAEA"
C_BLUE   = "#1B3A5C"; C_BLUE_BG   = "#E6EEF6"
C_BROWN  = "#5C4A1A"

def color_estat(estat):
    e = estat.lower()
    if "complet" in e:
        return C_GREEN, C_GREEN_BG
    if "progr" in e:
        return C_ORANGE, C_ORANGE_BG
    return C_GRAY, C_GRAY_BG

def color_prior(prior):
    p = prior.lower()
    if "alta" in p:
        return C_RED, C_RED_BG
    if "mitja" in p:
        return C_ORANGE, C_ORANGE_BG
    return C_GREEN, C_GREEN_BG

def color_risc(risc):
    r = risc.lower()
    if "alta" in r:
        return C_RED, C_RED_BG
    if "mitja" in r or "mig" in r:
        return C_ORANGE, C_ORANGE_BG
    return C_GREEN, C_GREEN_BG

def badge(text, fg, bg):
    s = "font-size:0.7rem;font-weight:500;padding:0.2rem 0.65rem;"
    s += "border-radius:2rem;white-space:nowrap;"
    s += "background:" + bg + ";color:" + fg
    return "<span style='" + s + "'>" + str(text) + "</span>"

def barra(pct, color):
    p = int(min(100, max(0, to_pct(pct))))
    s1 = "height:5px;background:#E2D9CE;border-radius:3px;overflow:hidden;margin-top:0.4rem"
    s2 = "width:" + str(p) + "%;height:100%;background:" + color + ";border-radius:3px"
    s3 = "font-size:0.72rem;color:#8C7B68;text-align:right;margin-top:2px"
    return "<div style='" + s1 + "'><div style='" + s2 + "'></div></div><div style='" + s3 + "'>" + str(p) + "%</div>"

def html_fases(ws, c_titol, c_desc, c_pct, c_estat):
    parts = []
    s_wrap = "display:flex;align-items:center;gap:1rem;padding:0.9rem 1.1rem;background:#FEFCF9;border:1px solid #E2D9CE;border-radius:0.75rem;margin-bottom:0.6rem"
    for r in files_dades(ws):
        titol = str(cel(ws, r, c_titol))
        desc  = str(cel(ws, r, c_desc))
        pct   = to_pct(cel(ws, r, c_pct))
        estat = str(cel(ws, r, c_estat))
        fg, bg = color_estat(estat)
        info  = "<div style='flex:1'><strong style='font-size:0.88rem;display:block'>" + titol + "</strong>"
        info += "<span style='font-size:0.76rem;color:#8C7B68'>" + desc + "</span></div>"
        b     = badge(estat, fg, bg)
        bar   = "<div style='width:80px'>" + barra(pct, fg) + "</div>"
        parts.append("<div style='" + s_wrap + "'>" + info + b + bar + "</div>")
    if not parts:
        return "<p style='color:#8C7B68'>Sense dades.</p>"
    return "\n".join(parts)

def html_tasques(ws, c_titol, c_fase, c_prior, c_estat, c_pct, c_notes):
    parts = []
    s_row = "display:flex;gap:0.75rem;padding:0.6rem 0.4rem;border-bottom:1px solid #E2D9CE"
    for r in files_dades(ws):
        titol = str(cel(ws, r, c_titol))
        fase  = str(cel(ws, r, c_fase))
        prior = str(cel(ws, r, c_prior))
        estat = str(cel(ws, r, c_estat))
        notes = str(cel(ws, r, c_notes))
        done  = "complet" in estat.lower()
        fg_p, bg_p = color_prior(prior)
        if done:
            style_t = "font-size:0.86rem;text-decoration:line-through;color:#8C7B68"
        else:
            style_t = "font-size:0.86rem;color:#1A1208"
        check  = "OK" if done else "[ ]"
        inner  = "<span style='font-size:0.85rem;margin-top:1px'>" + check + "</span>"
        inner += "<div style='flex:1'>"
        inner += "<div style='" + style_t + "'>" + titol + "</div>"
        inner += "<div style='display:flex;gap:0.6rem;margin-top:0.15rem;flex-wrap:wrap'>"
        inner += "<span style='font-size:0.72rem;color:#8C7B68'>" + fase + "</span>"
        inner += badge(prior, fg_p, bg_p)
        if notes:
            inner += "<span style='font-size:0.72rem;color:#8C7B68'>" + notes + "</span>"
        inner += "</div></div>"
        parts.append("<div style='" + s_row + "'>" + inner + "</div>")
    if not parts:
        return "<p style='color:#8C7B68'>Sense dades.</p>"
    return "\n".join(parts)

def html_riscos(ws, c_risc, c_prob, c_imp, c_mitig, c_estat):
    parts = []
    s_row = "display:flex;gap:0.75rem;padding:0.65rem 0;border-bottom:1px solid #E2D9CE;font-size:0.85rem"
    for r in files_dades(ws):
        risc  = str(cel(ws, r, c_risc))
        prob  = str(cel(ws, r, c_prob))
        mitig = str(cel(ws, r, c_mitig))
        fg, bg = color_risc(prob)
        inner  = badge(prob, fg, bg)
        inner += "<div><div style='color:#4A3F2F'>" + risc + "</div>"
        inner += "<div style='font-size:0.75rem;color:#8C7B68'>Mitigacio: " + mitig + "</div></div>"
        parts.append("<div style='" + s_row + "'>" + inner + "</div>")
    if not parts:
        return "<p style='color:#8C7B68'>Sense dades.</p>"
    return "\n".join(parts)

def html_pressupost(ws, c_conc, c_cat, c_est, c_notes):
    parts = []
    total = 0.0
    s_row = "display:flex;justify-content:space-between;align-items:center;padding:0.5rem 0;border-bottom:1px solid #E2D9CE;font-size:0.85rem"
    for r in files_dades(ws):
        conc = str(cel(ws, r, c_conc))
        cat  = str(cel(ws, r, c_cat)) if c_cat else ""
        est  = cel(ws, r, c_est)
        try:
            v = float(est)
            total += v
            est_txt = str(int(v)) + " EUR"
        except Exception:
            est_txt = "-"
        left  = "<span style='color:#4A3F2F'>" + conc
        left += "<span style='font-size:0.68rem;color:#8C7B68;background:#FAF7F2;padding:0.1rem 0.5rem;border-radius:2rem;margin-left:0.5rem'>" + cat + "</span></span>"
        right = "<span style='color:" + C_BROWN + "'>" + est_txt + "</span>"
        parts.append("<div style='" + s_row + "'>" + left + right + "</div>")
    s_tot = "display:flex;justify-content:space-between;padding:0.65rem 0;border-top:2px solid #E2D9CE;margin-top:0.25rem;font-weight:500;font-size:0.9rem"
    parts.append("<div style='" + s_tot + "'><span>Total estimat</span><span style='color:" + C_BROWN + ";font-size:1.05rem'>" + str(int(total)) + " EUR</span></div>")
    return "\n".join(parts)

def html_bibliografia(ws):
    parts = []
    s_row = "display:flex;gap:0.75rem;padding:0.65rem 0;border-bottom:1px solid #E2D9CE"
    for r in files_dades(ws):
        any_   = cel(ws, r, 2)
        ref    = str(cel(ws, r, 3))
        tema   = str(cel(ws, r, 4))
        llegit = str(cel(ws, r, 5))
        citat  = str(cel(ws, r, 6))
        if not ref or "Afegeix" in ref:
            continue
        dot = "<div style='width:6px;height:6px;min-width:6px;border-radius:50%;background:#3A5C8B;margin-top:0.55rem'></div>"
        any_txt = "(" + str(any_) + ") " if any_ else ""
        if "s" in llegit.lower():
            llegit_html = "<span style='font-size:0.68rem;color:#4A7C59'>Llegit</span>"
        else:
            llegit_html = "<span style='font-size:0.68rem;color:#8C7B68'>Pendent</span>"
        if "s" in citat.lower():
            citat_html = "<span style='font-size:0.68rem;color:#3A5C8B'>Citat</span>"
        else:
            citat_html = ""
        inner  = "<div style='font-size:0.85rem;color:#4A3F2F;line-height:1.6'>" + any_txt + ref + "</div>"
        inner += "<div style='font-size:0.72rem;color:#8C7B68;margin-top:0.1rem'>" + tema + " " + llegit_html + " " + citat_html + "</div>"
        parts.append("<div style='" + s_row + "'>" + dot + "<div>" + inner + "</div></div>")
    if not parts:
        return "<p style='color:#8C7B68'>Sense references.</p>"
    return "\n".join(parts)

def html_gramatica(ws):
    parts = []
    s_row = "display:flex;gap:0.75rem;padding:0.6rem 0.4rem;border-bottom:1px solid #E2D9CE"
    for r in files_dades(ws):
        tema  = str(cel(ws, r, 2))
        desc  = str(cel(ws, r, 3))
        fase  = str(cel(ws, r, 4))
        estat = str(cel(ws, r, 5))
        done  = "complet" in estat.lower()
        fg, bg = color_estat(estat)
        if done:
            style_t = "font-size:0.86rem;text-decoration:line-through;color:#8C7B68"
        else:
            style_t = "font-size:0.86rem;color:#1A1208"
        check  = "OK" if done else "[ ]"
        inner  = "<span style='font-size:0.85rem;margin-top:1px'>" + check + "</span>"
        inner += "<div style='flex:1'><div style='" + style_t + "'>" + tema + "</div>"
        inner += "<div style='display:flex;gap:0.6rem;margin-top:0.15rem;flex-wrap:wrap'>"
        inner += "<span style='font-size:0.72rem;color:#8C7B68;font-style:italic'>" + desc + "</span>"
        inner += badge(fase, C_BLUE, C_BLUE_BG)
        inner += "</div></div>"
        inner += badge(estat, fg, bg)
        parts.append("<div style='" + s_row + "'>" + inner + "</div>")
    if not parts:
        return "<p style='color:#8C7B68'>Sense dades.</p>"
    return "\n".join(parts)

def html_vocabulari(ws):
    parts = []
    s_row = "display:flex;align-items:center;gap:0.75rem;padding:0.55rem 0.4rem;border-bottom:1px solid #E2D9CE;font-size:0.85rem"
    for r in files_dades(ws):
        fr    = str(cel(ws, r, 2))
        cat   = str(cel(ws, r, 3))
        ex    = str(cel(ws, r, 4))
        categ = str(cel(ws, r, 5))
        estat = str(cel(ws, r, 6))
        if not fr.strip():
            continue
        fg, bg = color_estat(estat)
        inner  = "<div style='flex:1'>"
        inner += "<span style='font-weight:500;color:" + C_BLUE + "'>" + fr + "</span>"
        inner += "<span style='color:#8C7B68;margin:0 0.4rem'>-&gt;</span>"
        inner += "<span style='color:#4A3F2F'>" + cat + "</span>"
        inner += "<div style='font-size:0.72rem;color:#8C7B68;font-style:italic;margin-top:1px'>" + ex + "</div>"
        inner += "</div>"
        inner += badge(categ, "#4A3F2F", "#F0EEEB")
        inner += badge(estat, fg, bg)
        parts.append("<div style='" + s_row + "'>" + inner + "</div>")
    if not parts:
        return "<p style='color:#8C7B68'>Sense dades.</p>"
    return "\n".join(parts)

def html_registre(ws, max_files=10):
    parts = []
    totes    = files_dades(ws)
    sessions = [r for r in totes if cel(ws, r, 2) and str(cel(ws, r, 2)).strip()]
    sessions = sessions[-max_files:]
    s_bloc = "padding:0.9rem 0;border-bottom:1px solid #E2D9CE"
    s_cap  = "display:flex;justify-content:space-between;align-items:center;margin-bottom:0.4rem"
    for r in reversed(sessions):
        data  = str(cel(ws, r, 2))
        dia   = str(cel(ws, r, 3))
        hores = cel(ws, r, 4)
        acts  = str(cel(ws, r, 5))
        vocab = str(cel(ws, r, 6))
        obs   = str(cel(ws, r, 7))
        try:
            h_txt = str(round(float(hores), 1)) + "h"
        except Exception:
            h_txt = "-"
        cap  = "<div style='" + s_cap + "'>"
        cap += "<div><span style='font-weight:500;font-size:0.88rem;color:#1A1208'>" + data + "</span>"
        cap += "<span style='font-size:0.75rem;color:#8C7B68;margin-left:0.5rem'>" + dia + "</span></div>"
        cap += "<span style='color:" + C_BLUE + ";font-size:1rem'>" + h_txt + "</span></div>"
        body = "<div style='font-size:0.82rem;color:#4A3F2F'>" + acts + "</div>"
        if vocab:
            body += "<div style='font-size:0.78rem;color:#4A7C59;margin-top:0.2rem'>" + vocab + "</div>"
        if obs:
            body += "<div style='font-size:0.75rem;color:#8C7B68;font-style:italic;margin-top:0.2rem'>" + obs + "</div>"
        parts.append("<div style='" + s_bloc + "'>" + cap + body + "</div>")
    if not parts:
        return "<p style='color:#8C7B68'>Encara no hi ha sessions registrades.</p>"
    return "\n".join(parts)

def injectar(html_path, seccions):
    text = html_path.read_text(encoding="utf-8")
    for clau, contingut in seccions.items():
        pat = r'<!--\s*PLM:' + clau + r'\s*-->.*?<!--\s*/PLM:' + clau + r'\s*-->'
        rep = '<!-- PLM:' + clau + ' -->\n' + contingut + '\n<!-- /PLM:' + clau + ' -->'
        text, n = re.subn(pat, rep, text, flags=re.DOTALL)
        if n == 0:
            print("  avis: PLM:" + clau + " no trobat a " + html_path.name)
    html_path.write_text(text, encoding="utf-8")


def processar_tfm():
    xlsx = BASE / "tfm_quantica" / "PLM_TFM_Quantica.xlsx"
    html = BASE / "tfm_quantica" / "plm_tesi.html"
    if not xlsx.exists():
        print("  Excel tfm no trobat"); return
    wb   = load_workbook(xlsx, data_only=True)
    ws_c = nom_full(wb, ["Capitols", "Cap\u00edtols"])
    ws_t = wb["Tasques"]
    ws_r = wb["Riscos"]
    ws_b = wb["Bibliografia"]
    pg   = progress_global(ws_c, 6)
    cd   = sum(1 for r in files_dades(ws_c) if cel(ws_c, r, 8) == "Completat")
    td   = sum(1 for r in files_dades(ws_t) if cel(ws_t, r, 5) == "Completat")
    tt   = len(files_dades(ws_t))
    rt   = sum(1 for r in files_dades(ws_b) if cel(ws_b, r, 3) and "Afegeix" not in str(cel(ws_b, r, 3)))
    s = {
        "PROGRESS_PCT": str(int(pg)) + "%",
        "PROGRESS_BAR": str(int(pg)),
        "H_PROGRESS":   str(int(pg)) + "%",
        "H_CAPS":       str(cd) + "/" + str(len(files_dades(ws_c))),
        "H_TASQUES":    str(td) + "/" + str(tt),
        "H_REFS":       str(rt),
        "FASES":        html_fases(ws_c, 2, 3, 6, 8),
        "TASQUES":      html_tasques(ws_t, 2, 3, 4, 5, 6, 7),
        "RISCOS":       html_riscos(ws_r, 2, 3, 4, 5, 7),
        "BIBLIOGRAFIA": html_bibliografia(ws_b),
    }
    injectar(html, s)
    print("  OK  plm_tesi.html -> " + str(int(pg)) + "%")

def processar_elevador():
    xlsx = BASE / "elevador_plats" / "PLM_Elevador_Plats.xlsx"
    html = BASE / "elevador_plats" / "plm_elevador.html"
    if not xlsx.exists():
        print("  Excel elevador no trobat"); return
    wb   = load_workbook(xlsx, data_only=True)
    ws_f = wb["Fases"]
    ws_t = wb["Tasques"]
    ws_r = wb["Riscos"]
    ws_p = wb["Pressupost"]
    pg   = progress_global(ws_f, 6)
    fd   = sum(1 for r in files_dades(ws_f) if cel(ws_f, r, 7) == "Completat")
    td   = sum(1 for r in files_dades(ws_t) if cel(ws_t, r, 5) == "Completat")
    tt   = len(files_dades(ws_t))
    cost = 0.0
    for r in files_dades(ws_p):
        try:
            v = float(cel(ws_p, r, 6))
            if "TOTAL" not in str(cel(ws_p, r, 2)):
                cost += v
        except Exception:
            pass
    s = {
        "PROGRESS_PCT": str(int(pg)) + "%",
        "PROGRESS_BAR": str(int(pg)),
        "H_PROGRESS":   str(int(pg)) + "%",
        "H_FASES":      str(fd) + "/" + str(len(files_dades(ws_f))),
        "H_TASQUES":    str(td) + "/" + str(tt),
        "H_COST":       str(int(cost)) + " EUR" if cost > 0 else "-",
        "FASES":        html_fases(ws_f, 2, 3, 6, 7),
        "TASQUES":      html_tasques(ws_t, 2, 3, 4, 5, 6, 7),
        "RISCOS":       html_riscos(ws_r, 2, 3, 4, 5, 7),
        "PRESSUPOST":   html_pressupost(ws_p, 2, 3, 6, 8),
    }
    injectar(html, s)
    print("  OK  plm_elevador.html -> " + str(int(pg)) + "%")

def actualitzar_index():
    index = BASE / "index.html"
    if not index.exists():
        return
    configs = [
        ("tfm",  BASE / "tfm_quantica"   / "PLM_TFM_Quantica.xlsx",   ["Capitols", "Cap\u00edtols"],   6),
        ("elev", BASE / "elevador_plats" / "PLM_Elevador_Plats.xlsx", ["Fases"],                       6),
    ]
    prog = {}
    for pid, path, noms, c in configs:
        try:
            wb = load_workbook(path, data_only=True)
            ws = nom_full(wb, noms)
            prog[pid] = progress_global(ws, c)
        except Exception:
            prog[pid] = 0
    s = {
        "PCT_TFM":  str(int(prog["tfm"]))  + "%",
        "BAR_TFM":  str(int(prog["tfm"])),
        "PCT_ELEV": str(int(prog["elev"])) + "%",
        "BAR_ELEV": str(int(prog["elev"])),
    }
    injectar(index, s)
    print("  OK  index.html -> tfm " + str(int(prog["tfm"])) + "% | elev " + str(int(prog["elev"])) + "%")

def main():
    print("\n=== Actualitzant portafoli ===\n")
    print("\n[TFM Quantica]");     processar_tfm()
    print("\n[Elevador]");         processar_elevador()
    print("\n[Pagina principal]"); actualitzar_index()
    print("\nTot actualitzat.\n")

if __name__ == "__main__":
    main()
