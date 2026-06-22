#!/usr/bin/env python3
"""
actualitzar.py
Llegeix els tres Excel PLM i genera data.json a cada subcarpeta.
Executa'l des de la carpeta 'portafoli/' amb:  python actualitzar.py
"""

import json
import os
from pathlib import Path
from openpyxl import load_workbook

BASE = Path(__file__).parent

PROJECTES = [
    {
        "id": "web_personal",
        "xlsx": BASE / "web_personal" / "PLM_Web_Personal.xlsx",
        "json": BASE / "web_personal" / "data.json",
        "full_fases":      ("Fases",      2, 3, 5, 6, 8),   # titol, desc, data_fi, pct, estat
        "full_tasques":    ("Tasques",    2, 3, 4, 5, 6, 7), # titol, fase, prior, estat, pct, notes
        "full_riscos":     ("Riscos",     2, 3, 4, 5, 7),    # risc, prob, impacte, mitig, estat
        "full_pressupost": ("Pressupost", 2, 3, 4, 5, 6, 7), # concepte, cat, est, real, estat, notes
        "full_extra":      None,
    },
    {
        "id": "tfm_quantica",
        "xlsx": BASE / "tfm_quantica" / "PLM_TFM_Quantica.xlsx",
        "json": BASE / "tfm_quantica" / "data.json",
        "full_fases":      ("Capítols",    2, 3, 5, 6, 8),
        "full_tasques":    ("Tasques",     2, 3, 4, 5, 6, 7),
        "full_riscos":     ("Riscos",      2, 3, 4, 5, 7),
        "full_pressupost": None,
        "full_extra":      ("Bibliografia", [2, 3, 4, 5, 6, 7], ["any","referencia","tematica","llegit","citat","notes"]),
    },
    {
        "id": "elevador_plats",
        "xlsx": BASE / "elevador_plats" / "PLM_Elevador_Plats.xlsx",
        "json": BASE / "elevador_plats" / "data.json",
        "full_fases":      ("Fases",      2, 3, 5, 6, 7),
        "full_tasques":    ("Tasques",    2, 3, 4, 5, 6, 7),
        "full_riscos":     ("Riscos",     2, 3, 4, 5, 7),
        "full_pressupost": ("Pressupost", 2, 3, 6, 7, None, 8), # concepte, cat, cost_tot, cost_real, -, notes
        "full_extra":      None,
    },
]

def val(ws, row, col):
    """Llegeix el valor d'una cel·la, ignorant fórmules."""
    if col is None:
        return ""
    v = ws.cell(row=row, column=col).value
    if v is None:
        return ""
    if isinstance(v, float):
        return int(v) if v == int(v) else round(v, 2)
    if isinstance(v, str) and v.startswith("="):
        return ""  # Fórmula no avaluada — retornem buit
    return v

def es_fila_dades(ws, row_idx):
    """Comprova que la primera columna sigui un número de fila."""
    v = ws.cell(row=row_idx, column=1).value
    if v is None:
        return False
    try:
        int(str(v).strip())
        return True
    except (ValueError, AttributeError):
        return False

def llegir_fases(wb, nom, c_titol, c_desc, c_data, c_pct, c_estat):
    if nom not in wb.sheetnames:
        return []
    ws = wb[nom]
    files = []
    for r in range(5, ws.max_row + 1):
        if not es_fila_dades(ws, r):
            continue
        pct_raw = val(ws, r, c_pct)
        try:
            pct_v = float(pct_raw)
        except (ValueError, TypeError):
            pct_v = 0
        files.append({
            "titol":  val(ws, r, c_titol),
            "desc":   val(ws, r, c_desc),
            "data":   str(val(ws, r, c_data)),
            "pct":    pct_v,
            "estat":  val(ws, r, c_estat),
        })
    return files

def llegir_tasques(wb, nom, c_titol, c_fase, c_prior, c_estat, c_pct, c_notes):
    if nom not in wb.sheetnames:
        return []
    ws = wb[nom]
    files = []
    for r in range(5, ws.max_row + 1):
        if not es_fila_dades(ws, r):
            continue
        pct_raw = val(ws, r, c_pct)
        try:
            pct_v = float(pct_raw)
        except (ValueError, TypeError):
            pct_v = 0
        t = val(ws, r, c_titol)
        if not t:
            continue
        files.append({
            "titol":    t,
            "fase":     val(ws, r, c_fase),
            "prioritat":val(ws, r, c_prior),
            "estat":    val(ws, r, c_estat),
            "pct":      pct_v,
            "notes":    val(ws, r, c_notes),
        })
    return files

def llegir_riscos(wb, nom, c_risc, c_prob, c_imp, c_mitig, c_estat):
    if nom not in wb.sheetnames:
        return []
    ws = wb[nom]
    files = []
    for r in range(5, ws.max_row + 1):
        if not es_fila_dades(ws, r):
            continue
        risc = val(ws, r, c_risc)
        if not risc:
            continue
        files.append({
            "risc":     risc,
            "prob":     val(ws, r, c_prob),
            "impacte":  val(ws, r, c_imp),
            "mitigacio":val(ws, r, c_mitig),
            "estat":    val(ws, r, c_estat),
        })
    return files

def llegir_pressupost(wb, nom, c_conc, c_cat, c_est, c_real, c_estat, c_notes):
    if nom is None or nom not in wb.sheetnames:
        return []
    ws = wb[nom]
    files = []
    for r in range(5, ws.max_row + 1):
        if not es_fila_dades(ws, r):
            continue
        conc = val(ws, r, c_conc)
        if not conc:
            continue
        def num(c):
            if c is None: return 0
            v = val(ws, r, c)
            try: return float(v)
            except: return 0
        files.append({
            "concepte": conc,
            "categoria":val(ws, r, c_cat) if c_cat else "",
            "est":      num(c_est),
            "real":     num(c_real),
            "estat":    val(ws, r, c_estat) if c_estat else "",
            "notes":    val(ws, r, c_notes) if c_notes else "",
        })
    return files

def llegir_extra(wb, nom, cols, claus):
    if nom is None or nom not in wb.sheetnames:
        return []
    ws = wb[nom]
    files = []
    for r in range(5, ws.max_row + 1):
        if not es_fila_dades(ws, r):
            continue
        entrada = {}
        for clau, c in zip(claus, cols):
            entrada[clau] = val(ws, r, c)
        if any(v for v in entrada.values()):
            files.append(entrada)
    return files

def calcular_progress(fases):
    pcts = [f["pct"] for f in fases if isinstance(f["pct"], (int, float))]
    if not pcts:
        return 0
    return round(sum(pcts) / len(pcts), 1)

def processar(proj):
    if not proj["xlsx"].exists():
        print(f"  ⚠  No trobo {proj['xlsx']}")
        return

    wb = load_workbook(proj["xlsx"], data_only=True)

    ff = proj["full_fases"]
    ft = proj["full_tasques"]
    fr = proj["full_riscos"]
    fp = proj["full_pressupost"]
    fe = proj["full_extra"]

    fases      = llegir_fases(wb, ff[0], ff[1], ff[2], ff[3], ff[4], ff[5])
    tasques    = llegir_tasques(wb, ft[0], ft[1], ft[2], ft[3], ft[4], ft[5], ft[6])
    riscos     = llegir_riscos(wb, fr[0], fr[1], fr[2], fr[3], fr[4], fr[5])
    pressupost = llegir_pressupost(wb, fp[0] if fp else None,
                                   fp[1] if fp else None, fp[2] if fp else None,
                                   fp[3] if fp else None, fp[4] if fp else None,
                                   fp[5] if fp else None, fp[6] if fp else None) if fp else []
    extra      = llegir_extra(wb, fe[0], fe[1], fe[2]) if fe else []

    pg = calcular_progress(fases)

    dades = {
        "progress_global": pg,
        "fases":      fases,
        "tasques":    tasques,
        "riscos":     riscos,
        "pressupost": pressupost,
    }
    if fe:
        dades["bibliografia"] = extra

    with open(proj["json"], "w", encoding="utf-8") as f:
        json.dump(dades, f, ensure_ascii=False, indent=2, default=str)

    n_fases_done = sum(1 for f in fases if f["estat"] == "Completat")
    n_task_done  = sum(1 for t in tasques if t["estat"] == "Completat")
    print(f"  ✓  {proj['id']}/data.json")
    print(f"     Progrés: {pg}% | Fases: {n_fases_done}/{len(fases)} | Tasques: {n_task_done}/{len(tasques)}")

def main():
    print("\n=== Actualitzant dades dels projectes ===\n")
    for proj in PROJECTES:
        print(f"[{proj['id']}]")
        processar(proj)
    print("\n✓ Tots els data.json actualitzats.\n")

if __name__ == "__main__":
    main()
